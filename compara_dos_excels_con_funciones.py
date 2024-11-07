# Importamos las librerías necesarias
# pip install openpyxl 
import pandas as pd
import os
from openpyxl import load_workbook, Workbook

# Función principal para comparar y actualizar las diferencias acumuladas
def comparar_archivos(presupuesto, archivo_erp, nombre_hoja, diferencia_acumulada_ruta):
    try:
        # Carga los archivos Excel
        presupuesto_excel = f'ruta/{presupuesto}.xlsx'
        erp_excel = f'ruta/{archivo_erp}.xlsx'

        # Lee las hojas específicas en ambos archivos
        df_presupuesto = pd.read_excel(presupuesto_excel, sheet_name=nombre_hoja)
        df_costes = pd.read_excel(erp_excel, sheet_name=nombre_hoja)

        # Verificar si ambos archivos tienen las mismas columnas
        if df_presupuesto.columns.tolist() != df_costes.columns.tolist():
            print("Los archivos tienen diferentes columnas. Revísalos.")
            return

        # Establecemos una columna como índice para comparar
        df_presupuesto = df_presupuesto.set_index('ID')  # Cambia 'ID' según el nombre de la columna índice
        df_costes = df_costes.set_index('ID')

        # Encontrar diferencias entre los DataFrames
        diferencias = df_presupuesto.compare(df_costes, keep_shape=True, keep_equal=False)
        
        # Verificar si hay diferencias
        if diferencias.empty:
            print("Enhorabuena. Clavaste el presupuesto")
        else:
            print("Parece que ha habido diferencias. Revisa el Excel generado.")
            # Guardamos las diferencias en un nuevo archivo Excel
            excel_diferencias = f'ruta/{presupuesto}_Comparado.xlsx'
            diferencias.to_excel(excel_diferencias)

            # Guardamos o actualizamos el archivo de diferencias acumuladas
            actualizar_diferencias_acumuladas(diferencia_acumulada_ruta, diferencias)

    except FileNotFoundError as e:
        print(f"Error: No se encontró el archivo {e.filename}")
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")

# Función para actualizar las diferencias acumuladas en un archivo
def actualizar_diferencias_acumuladas(diferencias_acumuladas_ruta, diferencias):
    try:
        # Verifica si el archivo de diferencias acumuladas ya existe
        if os.path.exists(diferencias_acumuladas_ruta):
            # Si existe, carga el archivo y selecciona la hoja
            libro = load_workbook(diferencias_acumuladas_ruta)
            hoja = libro.active
        else:
            # Si no existe, crea un nuevo archivo y añade encabezados
            libro = Workbook()
            hoja = libro.active
            hoja.append(diferencias.columns.tolist())  # Agregar encabezados

        # Agrega las filas de diferencias al final de la hoja
        for index, row in diferencias.iterrows():
            hoja.append([index] + row.tolist())

        # Guarda el archivo (crea uno nuevo o actualiza el existente)
        libro.save(diferencias_acumuladas_ruta)

    except Exception as e:
        print(f"Error al actualizar las diferencias acumuladas: {e}")

# Especifica los nombres de archivo y ruta de salida
presupuesto = 'Presupuesto'  # Nombre del archivo de presupuesto
archivo_erp = 'archivo_erp'  # Nombre del archivo ERP
nombre_hoja = 'Hoja1'  # Nombre de la hoja
diferencias_acumuladas_ruta = 'ruta/al/archivo_acumulado.xlsx'  # Ruta del archivo acumulado

# Ejecuta la comparación de archivos
comparar_archivos(presupuesto, archivo_erp, nombre_hoja, diferencias_acumuladas_ruta)